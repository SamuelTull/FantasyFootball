import json
import pandas as pd
import random
import numpy as np
from difflib import get_close_matches
from openpyxl import load_workbook
import unicodedata


def json_normalise(df, col, add_prefix=False):
    # takes in a DataFrame and does json normalise on the column cols/ column in cols
    df = df.reset_index(drop=True)
    df_nested = df[col].apply(pd.Series)
    if add_prefix:
        df_nested = df_nested.add_prefix(col + "_")
    df = df.merge(df_nested, left_index=True, right_index=True)
    df = df.drop(columns=col)
    return df


def strip_string(s):
    return (
        "".join(
            c
            for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        )
        .replace(" ", "")
        .upper()
    )


def find_match(name):
    # check capitals, surname, umlauts etc
    potential = []
    x = strip_string(name)
    for i in df.index:
        y = strip_string(df["name"][i])
        if x == y:
            return [df["name"][i]]
        elif x in y:
            potential.append(df["name"][i])
    if len(potential) > 0:
        return potential
    else:
        return find_close_match(name)


def find_close_match(name):
    if " " in name:
        closeMatch = get_close_matches(name.title(), df["name"].values, n=3, cutoff=0.8)
    else:
        closeMatch = get_close_matches(
            name.title(), [x.split(" ", 1)[-1] for x in df["name"]], n=3, cutoff=0.9
        )
        if closeMatch:
            for name in df["name"]:
                for i in range(len(closeMatch)):
                    if closeMatch[i] == name.split(" ", 1)[-1]:
                        closeMatch[i] = name
    return closeMatch


def checkDraftable(i, manager):
    player = df.loc[i]["name"]
    if not df.loc[i]["available"]:
        if df.loc[i]["eliminated"]:
            print(f"Cannot draft {player}, team eliminated")
        else:
            playerOwner = df.loc[i]["manager"]
            print(f"Cannot draft {player}, owned by {playerOwner}")
        return False

    playerPosition = df.loc[i]["position"]
    maxPosition = {"  GKP": 2, " DEF": 5, " MID": 5, "FWD": 3}[playerPosition]
    managerPositionCount = len(
        df.query("manager == @manager").query("position==@playerPosition")
    )
    if managerPositionCount >= maxPosition:
        print(
            f"Cannot draft {player} as this would exceed the number of {playerPosition} ({maxPosition})"
        )
        return False

    return True


def backup():
    backUps[-2].to_csv(path.replace(".xlsx", "_backup_backup.csv"), index=False)
    backUps[-1].to_csv(path.replace(".xlsx", "_backup.csv"), index=False)
    backUps.append(df.copy())


def updateSheet(book, writer, dfSheet, sheetName):
    book[sheetName].delete_cols(1, 10)
    book[sheetName].delete_rows(1, len(df) + 1)
    dfSheet.to_excel(writer, sheet_name=sheetName, index=False)


def updateTables(draftNumber, managers=[], positions=[]):
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    for manager in managers:
        dfManager = df.query("manager == @manager")[
            cols_save + ["draftRound"]
        ].sort_values(["position", "draftRound"])
        updateSheet(book, writer, dfManager, manager)

    dfAvailable = df.query("available")[cols_save]
    updateSheet(book, writer, dfAvailable, "Available")

    for position in positions:
        updateSheet(
            book,
            writer,
            dfAvailable.query("position == @position"),
            "Available_" + position,
        )

    dfFuture = pd.DataFrame(
        {
            "draftNumber": np.arange(
                draftNumber + 2, min(squadSize * N + 1, draftNumber + 2 * N + 1)
            ),
            "manager": draftOrder[
                draftNumber + 1 : min(squadSize * N, draftNumber + 2 * N)
            ],
        }
    )
    dfHistory = (
        pd.concat([df.query("~available").query("~eliminated"), dfFuture])
        .sort_values("draftNumber", ascending=False)
        .fillna("")[cols_history]
    )
    dfHistory["turnsTillDraft"] = dfHistory["draftNumber"] - 2 - draftNumber
    updateSheet(book, writer, dfHistory, "History")

    writer.close()


def resetTables():
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    for manager in managers:
        dfManager = df.query("manager == @manager")[
            cols_save + ["draftRound"]
        ].sort_values(["position", "draftRound"])
        updateSheet(book, writer, dfManager, manager)

    dfAvailable = df.query("available")[cols_save]
    updateSheet(book, writer, dfAvailable, "Available")
    for position in ["  GKP", " DEF", " MID", "FWD"]:
        updateSheet(
            book,
            writer,
            dfAvailable.query("position == @position"),
            "Available_" + position,
        )

    dfFuture = pd.DataFrame(
        {
            "draftNumber": np.arange(1, 2 * N),
            "manager": draftOrder[0 : 2 * N - 1],
        }
    )
    dfHistory = (
        pd.concat([df.query("~available"), dfFuture])
        .sort_values("draftNumber", ascending=False)
        .fillna("")[cols_history]
    )
    dfHistory["turnsTillDraft"] = dfHistory["draftNumber"] - 1
    updateSheet(book, writer, dfHistory, "History")
    writer.close()


def resetRound(draftNumber):
    print("Resetting to start of draft round")
    df.available = backUps[-1].available
    df.manager = backUps[-1].manager
    df.draftNumber = backUps[-1].draftNumber
    df.draftRound = backUps[-1].draftRound
    for manager in managers:
        updateTables(draftNumber)
    for i in range(draftNumber - draftNumber % N, draftNumber + 1):
        manager = draftOrder[i]
        draft(i, manager)


def draft(draftNumber, manager):
    draftRound = draftNumber // N
    valid = False
    while not valid:
        player = input(f"\nRound {draftRound+1} Picking: {manager}... ")
        if player == "UNDO":
            resetRound(draftNumber)
            return
        print(f"Entered: {player}")
        if player in df["name"].values:
            valid = True
        else:
            potential = find_match(player)
            if len(potential) == 1:
                valid = True
                player = potential[0]
            elif len(potential) > 1:
                print(
                    "".join(potential[i] + ", " for i in range(len(potential)) if i < 4)
                )
                print("Which did you mean?")
            else:
                print(f"No player named: {player}")

        if valid:
            for i in df.index:
                if df.loc[i]["name"] == player:
                    if checkDraftable(i, manager):
                        df.at[i, "available"] = False
                        df.at[i, "manager"] = manager
                        df.at[i, "draftNumber"] = draftNumber + 1
                        df.at[i, "draftRound"] = draftRound + 1
                        playerPosition = df.at[i, "position"]
                    else:
                        valid = False
                    break
    print(f"{manager} drafted {player}")

    updateTables(draftNumber, [manager], [playerPosition])


def eliminateTeam(eliminatedTeams):
    for i in df.index:
        if df.loc[i]["team"] in eliminatedTeams:
            df.at[i, "eliminated"] = True
            df.at[i, "available"] = False
            df.at[i, "manager"] = "Eliminated"


if __name__ == "__main__":
    path = "/Users/Samuel Tull/OneDrive/Documents/NotCambridge/Draft22.xlsx"
    managers = [
        "Tull",
        "Oliver",
        "Openshaw",
        "Tomkinson",
        "Griffin",
        "Cornell",
        "Jarvis",
    ]
    random.shuffle(managers)
    N = len(managers)
    squadSize = 11
    draftOrder = []
    for i in range(squadSize):
        for j in range(N):
            if i % 2 == 0:
                draftOrder.append(managers[j])
            else:
                draftOrder.append(managers[N - 1 - j])

    cols_save = ["draftRank", "name", "team", "position"]
    cols_history = ["draftNumber", "manager", "name", "team", "position"]

    df = pd.read_json("players.json")
    df = pd.read_json("https://play.fifa.com/json/fantasy/players.json")

    squads = pd.read_json("squads.json")
    squads = pd.read_json("https://play.fifa.com/json/fantasy/squads_fifa.json")

    df = json_normalise(df, "stats")
    df = json_normalise(df, "matchDayPoints", True)
    squads["eliminated"] = squads["isActive"].apply(lambda x: not x)
    squads = squads.rename(columns={"id": "squadId", "name": "team"})
    squads = squads[["squadId", "team", "abbr", "seed", "group", "eliminated"]]
    df = df.merge(squads, on="squadId", how="left")
    df["cost"] = df["cost"].apply(lambda x: round(x / 1000000, 2))
    df["position"] = df["position"].map({1: "  GKP", 2: " DEF", 3: " MID", 4: "FWD"})
    df["available"] = True
    df["manager"] = None
    df["draftNumber"] = 0
    df["draftRound"] = 0
    minCost = df.groupby("position").min()["cost"].reset_index(name="minCost")
    df = df.merge(minCost, on="position", how="left")
    df["value"] = df["cost"] - df["minCost"]
    df = df.sort_values(["value", "pickedBy"], ascending=False)
    df = df.reset_index(drop=True)
    df["draftRank"] = df.index + 1

    eliminateTeam(["Qatar", "Ecuador"])
    resetTables()

    print(f"Saving data to: {path}")
    for i in range(50):
        print("")
    print("Welcome to the World Cup 2022 draft")
    print("Managers in draft order:")
    for i, manager in enumerate(managers):
        print(i + 1, manager)

    backUps = [df, df]
    for i in range(squadSize * N):
        if i % N == 0:
            backup()
        manager = draftOrder[i]
        draft(i, manager)

    df.to_csv(path.replace(".xlsx", "_completed.csv"), index=False)
    df.to_csv("Draft22Netflix_completed.csv", index=False)
