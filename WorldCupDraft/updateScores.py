import json
import pandas as pd
import random
import numpy as np
from difflib import get_close_matches
from openpyxl import load_workbook
import unicodedata


def json_normalise(df, col, add_prefix=False):
    # takes in a DataFrame and does json normalise on the column cols/ column in cols
    df = df.reset_index(drop=True)
    df_nested = df[col].apply(pd.Series)
    if add_prefix:
        df_nested = df_nested.add_prefix(col + "_")
    df = df.merge(df_nested, left_index=True, right_index=True)
    df = df.drop(columns=col)
    return df


def updateSheet(book, writer, dfSheet, sheetName):
    book[sheetName].delete_cols(1, 10)
    book[sheetName].delete_rows(1, len(df) + 1)
    dfSheet.to_excel(writer, sheet_name=sheetName, index=False)


def updateTables(draftNumber, managers=[], positions=[]):
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    for manager in managers:
        dfManager = df.query("manager == @manager")[
            cols_save + ["draftRound"] + [f"GW{i}" for i in range(1, currGW + 1)]
        ].sort_values(["position", "draftRound"])
        updateSheet(book, writer, dfManager, manager)

    dfAvailable = df.query("available")[cols_save]
    updateSheet(book, writer, dfAvailable, "Available")

    for position in positions:
        updateSheet(
            book,
            writer,
            dfAvailable.query("position == @position"),
            "Available_" + position,
        )

    dfFuture = pd.DataFrame(
        {
            "draftNumber": np.arange(
                draftNumber + 2, min(squadSize * N + 1, draftNumber + 2 * N + 1)
            ),
            "manager": draftOrder[
                draftNumber + 1 : min(squadSize * N, draftNumber + 2 * N)
            ],
        }
    )
    dfHistory = (
        pd.concat([df.query("~available").query("~eliminated"), dfFuture])
        .sort_values("draftNumber", ascending=False)
        .fillna("")[cols_history]
    )
    dfHistory["turnsTillDraft"] = dfHistory["draftNumber"] - 2 - draftNumber
    # updateSheet(book, writer, dfHistory, "History")

    writer.close()


def updateLeaderboard():
    listGWs = [f"GW{i}" for i in range(1, currGW + 1)]
    leaderboard = df.groupby("manager").sum()[listGWs].reset_index()
    leaderboard["Total"] = sum(leaderboard[x] for x in listGWs)
    leaderboard = leaderboard.query("manager in @managers")
    leaderboard = leaderboard.sort_values("Total", ascending=False)

    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    updateSheet(book, writer, leaderboard, "Table")
    writer.close()


def resetTables():
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    for manager in managers:
        dfManager = df.query("manager == @manager")[
            cols_save + ["draftRound"]
        ].sort_values(["position", "draftRound"])
        updateSheet(book, writer, dfManager, manager)

    dfAvailable = df.query("available")[cols_save]
    updateSheet(book, writer, dfAvailable, "Available")
    for position in ["  GKP", " DEF", " MID", "FWD"]:
        updateSheet(
            book,
            writer,
            dfAvailable.query("position == @position"),
            "Available_" + position,
        )

    dfFuture = pd.DataFrame(
        {
            "draftNumber": np.arange(1, 2 * N),
            "manager": draftOrder[0 : 2 * N - 1],
        }
    )
    dfHistory = (
        pd.concat([df.query("~available"), dfFuture])
        .sort_values("draftNumber", ascending=False)
        .fillna("")[cols_history]
    )
    dfHistory["turnsTillDraft"] = dfHistory["draftNumber"] - 1
    # updateSheet(book, writer, dfHistory, "History")
    writer.close()


def updateScores():
    gameweekDays = {}
    gameweekDays[1] = [
        "matchDayPoints_1",
        "matchDayPoints_2",
        "matchDayPoints_3",
        "matchDayPoints_4",
    ]
    gameweekDays[2] = [
        "matchDayPoints_5",
        "matchDayPoints_6",
        "matchDayPoints_7",
        "matchDayPoints_8",
    ]
    gameweekDays[3] = [
        "matchDayPoints_9",
        "matchDayPoints_10",
        "matchDayPoints_11",
        "matchDayPoints_12",
    ]
    gameweekDays[4] = [
        "matchDayPoints_13",
        "matchDayPoints_14",
        "matchDayPoints_15",
        "matchDayPoints_16",
    ]

    gameweekDays[5] = ["matchDayPoints_17", "matchDayPoints_18"]
    gameweekDays[6] = ["matchDayPoints_19", "matchDayPoints_20"]
    gameweekDays[7] = ["matchDayPoints_21", "matchDayPoints_22"]
    for week in range(1, 8):
        df[f"GW{week}"] = sum(df[i] for i in gameweekDays[week])


if __name__ == "__main__":
    for i in range(2):
        if i == 0:
            path = "/Users/Samuel Tull/OneDrive/Documents/NotCambridge/Draft22.xlsx"
            data_path = "Draft22Netflix_completed.csv"
            managers = [
                "Tull",
                "Oliver",
                "Openshaw",
                "Tomkinson",
                "Griffin",
                "Cornell",
                "Jarvis",
            ]
        else:
            path = "/Users/Samuel Tull/OneDrive/Documents/NotCambridge/DraftSamSam.xlsx"
            data_path = "Draft22_completed_SAMSAM.csv"
            managers = ["Tull", "Oliver"]
        currGW = 3
        N = len(managers)
        squadSize = 11

        draftOrder = []
        for i in range(squadSize):
            for j in range(N):
                if i % 2 == 0:
                    draftOrder.append(managers[j])
                else:
                    draftOrder.append(managers[N - 1 - j])
        cols_save = ["draftRank", "name", "team", "position"]
        cols_history = ["draftNumber", "manager", "name", "team", "position"]
        cols_df = [
            "id",
            "team",
            "abbr",
            "seed",
            "group",
            "eliminated",
            "available",
            "manager",
            "draftNumber",
            "draftRound",
            "draftRank",
        ]
        df = pd.read_json("https://play.fifa.com/json/fantasy/players.json")
        df = json_normalise(df, "stats")
        df = json_normalise(df, "matchDayPoints", True)
        df = df.fillna(0)
        df["position"] = df["position"].map(
            {1: "  GKP", 2: " DEF", 3: " MID", 4: "FWD"}
        )
        updateScores()
        df_complete = pd.read_csv(data_path)[cols_df]
        df = df_complete.merge(df, how="right", on="id")

        updateTables(N * squadSize - 1, managers, ["  GKP", " DEF", " MID", "FWD"])
        updateLeaderboard()
